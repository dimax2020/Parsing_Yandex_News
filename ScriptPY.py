#!/usr/bin/python3


from selenium import webdriver   # pip install selenium
import openpyxl                  # pip install openpyxl


driver = webdriver.Firefox(executable_path='/my/directory/geckodriver')    # Указываем путь к deckodriver.
driver.get("https://yandex.com/news/")                                      # Открывает страничку яндекс новостей.

wb = openpyxl.load_workbook('/my/directory/News.xlsx')    # Открывает документ Excel, его придется создать 
sheet = wb['Лист1']    # Указываем лист в Excel.            самому через MOffice или через LibreCalc.
sheet['A1'] = 'News'
sheet['B1'] = 'Href'   # Создает заголовки для удобства.
wb.save('/my/directory/News.xlsx')
rows = sheet.max_row
cols = sheet.max_column
print(rows, cols)      #Пишет количество строк и столбцов в файле Excel.

allnews = driver.find_elements_by_xpath('//a[@class="news-card__link"]')    # Ишем кнопку по классу с помощью xpath.
for k in allnews:
    rows = sheet.max_row
    a = 0
    for i in range(1, rows + 1):            # Далее идет проверка на совпадения, но когда в файле уже много новостей,
        cell = sheet.cell(row=i, column=1)  # это происходит довольно долго.
        if k.text == cell.value:
            a += 1
            print('еще' , a, 'совпадение')
    if a == 0:
        cell = sheet.cell(row=rows + 1, column=1)
        cell.value = k.text
        cell = sheet.cell(row=rows + 1, column=2)     # Заполнение таблицы.
        cell.value = k.get_attribute('href')
        wb.save('/my/directory/News.xlsx')

driver.quit()    # Автоматически закрывает браузер для удобства.
